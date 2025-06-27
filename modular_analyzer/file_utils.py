# --- modular_analyzer/file_utils.py ---

import csv
import logging
import os
import tempfile
import zipfile

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def list_yaml_configs(config_dir):
    return [f for f in os.listdir(config_dir) if f.endswith(".yaml")]


def load_yaml(path):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def save_csv(data, columns, filepath):
    if not data:
        logging.warning(f"No data to write for {filepath}")
        return
    with open(filepath, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows(data)


def color_code_excel(output_dir):
    try:
        filepath = os.path.join(output_dir, "ticket_numbers.csv")
        df = pd.read_csv(filepath)
        xlsx_path = filepath.replace(".csv", ".xlsx")
        df.to_excel(xlsx_path, index=False)

        wb = load_workbook(xlsx_path)
        ws = wb.active

        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and "missing" in cell.value.lower():
                    cell.fill = yellow_fill
                elif isinstance(cell.value, str) and "template" in cell.value.lower():
                    cell.fill = red_fill

        wb.save(xlsx_path)
        logging.info(f"Excel file saved with highlights: {xlsx_path}")
    except Exception as e:
        logging.error(f"Failed to color-code Excel: {e}")


def zip_folder(folder_path, output_zip_path):
    if not os.path.exists(folder_path):
        logging.warning(f"Zip target folder does not exist: {folder_path}")
        return
    with zipfile.ZipFile(output_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, _, files in os.walk(folder_path):
            for file in files:
                abs_path = os.path.join(root, file)
                rel_path = os.path.relpath(abs_path, folder_path)
                zipf.write(abs_path, rel_path)
    logging.info(f"Zipped valid pages: {output_zip_path}")


def find_file_case_insensitive(target_filename, search_dir):
    target_lower = target_filename.lower()
    for root, _, files in os.walk(search_dir):
        for file in files:
            if file.lower() == target_lower:
                return os.path.join(root, file)
    return None


def validate_required_files(vendor_name, config_dir, template_files, template_dir):
    missing = []

    yaml_found = (
            find_file_case_insensitive(f"{vendor_name}.yaml", config_dir) or
            find_file_case_insensitive(f"{vendor_name}.yml", config_dir)
    )
    if not yaml_found:
        missing.append(f"Vendor config: {vendor_name}.yaml/yml")

    for template_name in template_files:
        template_path = find_file_case_insensitive(template_name, template_dir)
        if not template_path:
            missing.append(f"Template: {template_name}")

    return (len(missing) == 0), missing


def is_dir_writable(path):
    try:
        with tempfile.TemporaryFile(dir=path):
            return True
    except Exception:
        return False


def save_entries_to_excel(entries, output_dir, base_name):
    """
    Save extracted ticket entries to an Excel file.
    """
    if not entries:
        logging.warning("No entries to write to Excel.")
        return

    df = pd.DataFrame(entries)
    csv_path = os.path.join(output_dir, f"{base_name}_ticket_numbers.csv")
    xlsx_path = csv_path.replace(".csv", ".xlsx")

    df.to_csv(csv_path, index=False)
    df.to_excel(xlsx_path, index=False)

    logging.info(f"Saved ticket data to:\n  CSV: {csv_path}\n  Excel: {xlsx_path}")
